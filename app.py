from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_from_directory, send_file
from werkzeug.exceptions import HTTPException
import sqlite3, os, json, secrets, hashlib, hmac, re, io, time, logging
from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.middleware.proxy_fix import ProxyFix
from PIL import Image, ImageOps, UnidentifiedImageError

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

try:
    Image.MAX_IMAGE_PIXELS = 25_000_000
except Exception:
    pass

app=Flask(__name__)

# Monetary values are stored as integer cents (USD).
def money(value):
    try:
        cents=int(value or 0)
    except (TypeError, ValueError):
        cents=0
    sign="-" if cents < 0 else ""
    cents=abs(cents)
    return f"{sign}${cents//100:,}.{cents%100:02d}"

def parse_money_cents(value):
    raw=str(value or "").strip().replace("$", "").replace(",", "")
    if not raw:
        return 0
    try:
        amount=float(raw)
    except (TypeError, ValueError):
        raise ValueError("Số tiền không hợp lệ.")
    if amount < 0:
        raise ValueError("Số tiền không được âm.")
    cents=int(round(amount*100))
    return cents

app.jinja_env.filters["money"] = money

# Load .env located beside app.py. No external dependency required.
def load_local_env():
    env_path=os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.isfile(env_path):
        return
    try:
        with open(env_path, "r", encoding="utf-8") as f:
            for raw in f:
                line=raw.strip()
                if not line or line.startswith("#") or "=" not in line:
                    continue
                key,value=line.split("=",1)
                key=key.strip()
                value=value.strip()
                if len(value)>=2 and value[0]==value[-1] and value[0] in ("'", '"'):
                    value=value[1:-1]
                os.environ.setdefault(key,value)
    except OSError:
        pass

load_local_env()

SECRET_KEY=os.getenv("SECRET_KEY","")
if not SECRET_KEY or SECRET_KEY.startswith(("REPLACE_","PASTE_")):
    raise RuntimeError("SECRET_KEY chưa được cấu hình trong file .env.")
app.secret_key=SECRET_KEY
app.config.update(
    MAX_CONTENT_LENGTH=8*1024*1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SECURE=os.getenv("SESSION_COOKIE_SECURE","0")=="1",
    SESSION_COOKIE_SAMESITE="Lax",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
)
BASE=os.path.dirname(__file__); DB=os.path.join(BASE,"data","kk.db"); UP=os.path.join(BASE,"uploads")
IMPORT_DIR=os.path.join(BASE,"data","import_staging")
MAX_IMPORT_ROWS=500
ADMIN_PASSWORD_HASH=os.getenv("ADMIN_PASSWORD_HASH","")
ADMIN_USERNAME=os.getenv("ADMIN_USERNAME","admin")
ADMIN_IDLE_TIMEOUT=30*60
TRUST_PROXY=os.getenv("TRUST_PROXY","0")=="1"
FORCE_HTTPS=os.getenv("FORCE_HTTPS","0")=="1"
HSTS_ENABLED=os.getenv("HSTS_ENABLED","0")=="1"
CSP_ENABLED=os.getenv("CSP_ENABLED","0")=="1"
MAX_IMAGE_PIXELS=25_000_000
RATE_LIMITS={
    "login": (5, 15*60),
    "order": (8, 10*60),
    "booking": (5, 10*60),
}
BACKUP_DIR=os.path.join(BASE if "BASE" in globals() else os.path.dirname(__file__), "backups")
LOG_DIR=os.path.join(BASE if "BASE" in globals() else os.path.dirname(__file__), "logs")
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(IMPORT_DIR, exist_ok=True)

if TRUST_PROXY:
    app.wsgi_app=ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

logger=logging.getLogger("kkrestaurant")
if not logger.handlers:
    logger.setLevel(logging.INFO)
    _handler=RotatingFileHandler(os.path.join(LOG_DIR,"app.log"), maxBytes=2*1024*1024, backupCount=5, encoding="utf-8")
    _handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    logger.addHandler(_handler)

CATEGORIES=[("all","Tất cả"),("pho","Phở"),("banh-mi","Bánh mì"),("bun","Bún"),("com","Cơm"),("chay","Món chay"),("trang-mieng","Tráng miệng")]
DEFAULT_CATEGORIES=[("pho","Phở"),("banh-mi","Bánh mì"),("bun","Bún"),("com","Cơm"),("chay","Món chay"),("trang-mieng","Tráng miệng")]
SEED_MENU=[{'id': 'pho-dac-biet', 'name': 'Phở Đặc Biệt', 'category': 'pho', 'price': 890, 'desc': 'Phở kết hợp bò tái, nạm, gân, bò viên và nước dùng thơm đậm vị.', 'badge': 'Bán chạy'}, {'id': 'pho-bo-tai', 'name': 'Phở Bò Tái', 'category': 'pho', 'price': 790, 'desc': 'Phở bò tái với nước dùng trong, thơm và đậm đà.'}, {'id': 'pho-tai-chin', 'name': 'Phở Tái Chín', 'category': 'pho', 'price': 790, 'desc': 'Phở bò tái và nạm chín.'}, {'id': 'pho-bo-vien', 'name': 'Phở Bò Viên', 'category': 'pho', 'price': 790, 'desc': 'Phở bò viên với nước dùng bò truyền thống.'}, {'id': 'pho-chin-vien', 'name': 'Phở Chín Viên', 'category': 'pho', 'price': 790, 'desc': 'Phở nạm chín và bò viên.'}, {'id': 'pho-ga', 'name': 'Phở Gà', 'category': 'pho', 'price': 790, 'desc': 'Phở gà với nước dùng nhẹ, thơm.'}, {'id': 'pho-tom', 'name': 'Phở Tôm', 'category': 'pho', 'price': 890, 'desc': 'Phở tôm, lựa chọn nước dùng phù hợp.'}, {'id': 'banh-mi-dac-biet', 'name': 'Bánh Mì Đặc Biệt', 'category': 'banh-mi', 'price': 790, 'desc': 'Bánh mì thịt heo, giăm bông Việt Nam và pate.', 'badge': 'Signature'}, {'id': 'banh-mi-thit-nuong', 'name': 'Bánh Mì Thịt Nướng', 'category': 'banh-mi', 'price': 690, 'desc': 'Bánh mì thịt heo nướng thơm.'}, {'id': 'banh-mi-ga', 'name': 'Bánh Mì Gà', 'category': 'banh-mi', 'price': 690, 'desc': 'Bánh mì gà nướng.'}, {'id': 'banh-mi-xa-xiu', 'name': 'Bánh Mì Xá Xíu', 'category': 'banh-mi', 'price': 690, 'desc': 'Bánh mì thịt heo xá xíu.'}, {'id': 'banh-mi-ca-ri-ga', 'name': 'Bánh Mì Cà Ri Gà', 'category': 'banh-mi', 'price': 690, 'desc': 'Gà cà ri dùng kèm bánh mì Pháp.'}, {'id': 'banh-mi-bo-kho', 'name': 'Bánh Mì Bò Kho', 'category': 'banh-mi', 'price': 790, 'desc': 'Bò kho Việt Nam dùng kèm bánh mì.'}, {'id': 'bun-cha', 'name': 'Bún Chả Giò Thịt/Tôm/Gà Nướng', 'category': 'bun', 'price': 890, 'desc': 'Bún với rau sống, dưa chua, giá, chả giò và lựa chọn thịt nướng, tôm hoặc gà.', 'badge': 'Bán chạy'}, {'id': 'bun-thit-ga-tom', 'name': 'Bún Thịt/Gà/Tôm Nướng', 'category': 'bun', 'price': 890, 'desc': 'Bún rau sống, bạc hà, dưa leo, cà rốt ngâm và đậu phộng.'}, {'id': 'bun-thit-nem', 'name': 'Bún Thịt Nem Nướng', 'category': 'bun', 'price': 890, 'desc': 'Thịt nướng và nem nướng dùng với bún.'}, {'id': 'bun-bo-xao-xa-ot', 'name': 'Bún Bò/Gà/Tôm Xào Xả Ớt', 'category': 'bun', 'price': 890, 'desc': 'Bún với bò, gà hoặc tôm xào sả ớt.'}, {'id': 'hu-tieu-mi', 'name': 'Hủ Tiếu Mì', 'category': 'bun', 'price': 890, 'desc': 'Hủ tiếu mì hải sản hoặc trứng.'}, {'id': 'com-bo-xao-luc-lac', 'name': 'Cơm Bò Xào Lúc Lắc', 'category': 'com', 'price': 990, 'desc': 'Bò xào bông cải, ớt chuông, hành tây và cơm trắng.', 'badge': 'Signature'}, {'id': 'com-bo-ga-tom-xao', 'name': 'Cơm Bò/Gà/Tôm Xào Sả Ớt', 'category': 'com', 'price': 990, 'desc': 'Bò, gà hoặc tôm xào sả ớt với cơm trắng.'}, {'id': 'com-ga-thit-tom-nuong', 'name': 'Cơm Gà/Thịt/Tôm Nướng', 'category': 'com', 'price': 990, 'desc': 'Cơm trắng, salad và lựa chọn món nướng.'}, {'id': 'com-tay-cam', 'name': 'Cơm Tay Cầm', 'category': 'com', 'price': 1090, 'desc': 'Cơm chiên kết hợp tôm, gà, bò xào rau củ trong thố nóng.'}, {'id': 'che-ba-mau', 'name': 'Chè Ba Màu', 'category': 'trang-mieng', 'price': 490, 'desc': 'Chè ba màu với thạch và nước cốt dừa.'}, {'id': 'kem-chuoi', 'name': 'Kem Chuối/Chiên', 'category': 'trang-mieng', 'price': 490, 'desc': 'Chuối chiên dùng kèm kem.'}, {'id': 'yogurt', 'name': 'Yogurt', 'category': 'trang-mieng', 'price': 390, 'desc': 'Sữa chua Việt Nam.'}, {'id': 'tofu-cuon-chay', 'name': 'Tofu Cuốn', 'category': 'chay', 'price': 690, 'desc': 'Đậu hũ chiên cuốn bánh tráng, rau sống và sốt đậu phộng.'}, {'id': 'banh-mi-chay', 'name': 'Bánh Mì Chay', 'category': 'chay', 'price': 590, 'desc': 'Đậu hũ, nấm xào cuốn bánh mì cùng rau và sốt.'}, {'id': 'goi-chay', 'name': 'Gỏi Chay', 'category': 'chay', 'price': 690, 'desc': 'Gỏi bắp cải với đậu hũ xào và đậu phộng.'}, {'id': 'pho-chay', 'name': 'Phở Chay', 'category': 'chay', 'price': 790, 'desc': 'Phở rau củ, đậu hũ, giá, húng quế và jalapeño.'}, {'id': 'bun-xao-chay', 'name': 'Bún Xào Chay', 'category': 'chay', 'price': 690, 'desc': 'Đậu hũ, bắp cải, cà rốt, mì gạo, dưa leo, rau xà lách và bạc hà.'}, {'id': 'mi-xao-chay', 'name': 'Mì Xào Chay', 'category': 'chay', 'price': 690, 'desc': 'Mì trứng xào đậu hũ và rau củ.'}, {'id': 'com-tay-cam-chay', 'name': 'Cơm Tay Cầm Chay', 'category': 'chay', 'price': 790, 'desc': 'Rau củ và đậu hũ xào trong thố nóng.'}]
DEFAULT_SETTINGS={"restaurant_name":"Phở & Bánh Mì K&K","tagline":"AUTHENTIC VIETNAMESE CUISINE","phone":"510.666.9966","address":"Cập nhật địa chỉ nhà hàng trong Admin","maps_url":"https://www.google.com/maps/search/?api=1&query=Pho+%26+Banh+Mi+K%26K+510.666.9966","google_review_url":"","hero_title":"Hương vị Việt, trọn vẹn trong từng món.","delivery_note":"Đặt món online nhanh chóng • Thanh toán khi nhận hàng (COD)","hero_image":"/static/menu-board.png","logo_image":"","online_order_enabled":"0","currency":"USD","currency_symbol":"$","location_eyebrow":"LOCATION","location_heading":"Địa điểm K&K","location_title":"Ghé K&K thưởng thức món ngon","location_description":"K&K phục vụ phở, bánh mì và các món Việt quen thuộc. Quý khách có thể xem thực đơn trước khi ghé nhà hàng và liên hệ trực tiếp với K&K nếu cần hỗ trợ.","location_main_image":"/static/menu-board.png","location_side_image_1":"/uploads/banh-mi-dac-biet_9353a2073b16.png","location_side_image_2":"/uploads/pho-dac-biet_d3d371fb4ca8.png"}

ORDER_STATUSES=["Mới","Đang chuẩn bị","Đang giao","Hoàn tất","Đã hủy"]
BOOKING_STATUSES=["Mới","Đã xác nhận","Hoàn tất","Đã hủy"]
VOUCHER_TYPES={"percent","fixed"}
IMAGE_EXTENSIONS={"jpg","jpeg","png","webp"}

def client_ip():
    """Return the client IP. Only trust X-Forwarded-For when TRUST_PROXY=1."""
    if TRUST_PROXY:
        return request.headers.get("X-Forwarded-For","").split(",")[0].strip() or request.remote_addr or "unknown"
    return request.remote_addr or "unknown"

def _rate_identity(scope, extra=""):
    raw=f"{scope}|{client_ip()}|{extra}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()

def consume_rate_limit(scope, extra=""):
    """SQLite-backed rate limit so it works across multiple Gunicorn/Waitress workers."""
    limit, window=RATE_LIMITS[scope]
    key=_rate_identity(scope, extra)
    now=int(time.time())
    c=conn()
    try:
        c.execute("DELETE FROM rate_limit_hits WHERE created_at < ?", (now-window,))
        count=c.execute(
            "SELECT COUNT(*) n FROM rate_limit_hits WHERE scope=? AND key_hash=? AND created_at>=?",
            (scope,key,now-window)
        ).fetchone()["n"]
        if count >= limit:
            c.commit()
            return False
        c.execute("INSERT INTO rate_limit_hits(scope,key_hash,created_at) VALUES(?,?,?)",(scope,key,now))
        c.commit()
        return True
    finally:
        c.close()

def audit_log(action, entity_type="", entity_id="", details=None):
    """Write an admin audit event without ever storing passwords or secrets."""
    try:
        username=str(session.get("admin_username") or "system")
        payload=json.dumps(details or {}, ensure_ascii=False, default=str)
        c=conn()
        c.execute("""INSERT INTO audit_logs(created_at,username,ip,action,entity_type,entity_id,details)
                    VALUES(?,?,?,?,?,?,?)""",
                  (datetime.now().strftime("%Y-%m-%d %H:%M:%S"),username,client_ip(),
                   action,entity_type,str(entity_id),payload[:4000]))
        c.commit(); c.close()
    except Exception:
        logger.exception("audit_log failed")

def save_validated_image(file_obj, path, requested_ext=None):
    """Validate real image content and re-encode it to strip dangerous payloads/metadata."""
    if not file_obj or not file_obj.filename:
        raise ValueError("Vui lòng chọn file ảnh.")
    ext=(requested_ext or "").lower().lstrip(".")
    if ext not in IMAGE_EXTENSIONS:
        raise ValueError("Ảnh phải có định dạng JPG, JPEG, PNG hoặc WEBP.")
    try:
        file_obj.stream.seek(0)
        with Image.open(file_obj.stream) as probe:
            actual=probe.format
            if actual not in {"JPEG","PNG","WEBP"}:
                raise ValueError("Nội dung file không phải ảnh JPG, PNG hoặc WEBP hợp lệ.")
            if actual.lower() != ("jpg" if ext=="jpeg" else ext):
                # Extension mismatch is allowed only when both are equivalent JPEG names.
                if not (actual=="JPEG" and ext in {"jpg","jpeg"}):
                    raise ValueError("Phần mở rộng không khớp với nội dung ảnh.")
            width,height=probe.size
            if width<1 or height<1 or width*height>MAX_IMAGE_PIXELS:
                raise ValueError("Kích thước ảnh vượt giới hạn an toàn.")
            probe.verify()
        file_obj.stream.seek(0)
        with Image.open(file_obj.stream) as img:
            img=ImageOps.exif_transpose(img)
            width,height=img.size
            if actual=="JPEG":
                if img.mode not in ("RGB","L"):
                    img=img.convert("RGB")
                img.save(path, format="JPEG", quality=90, optimize=True, progressive=True)
            elif actual=="PNG":
                if img.mode not in ("RGB","RGBA","L","LA","P"):
                    img=img.convert("RGBA")
                img.save(path, format="PNG", optimize=True)
            else:
                if img.mode not in ("RGB","RGBA"):
                    img=img.convert("RGBA" if "A" in img.mode else "RGB")
                img.save(path, format="WEBP", quality=90, method=6)
    except (UnidentifiedImageError, OSError, Image.DecompressionBombError) as exc:
        raise ValueError("File ảnh không hợp lệ hoặc bị hỏng.") from exc
    finally:
        try: file_obj.stream.seek(0)
        except Exception: pass

_backup_last_check=0.0
def backup_database(force=False):
    """Create a consistent SQLite backup using the SQLite backup API."""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp=datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    target=os.path.join(BACKUP_DIR,f"kk_{stamp}.db")
    src=None; dst=None
    try:
        src=sqlite3.connect(DB)
        dst=sqlite3.connect(target)
        src.backup(dst)
        dst.close(); dst=None
        src.close(); src=None
        logger.info("Database backup created: %s", target)
        # Keep the newest 14 backups.
        files=sorted([os.path.join(BACKUP_DIR,x) for x in os.listdir(BACKUP_DIR)
                      if x.startswith("kk_") and x.endswith(".db")], key=os.path.getmtime, reverse=True)
        for old in files[14:]:
            try: os.remove(old)
            except OSError: pass
        return target
    except Exception:
        logger.exception("Database backup failed")
        for obj in (dst,src):
            try:
                if obj: obj.close()
            except Exception: pass
        try:
            if os.path.exists(target) and os.path.getsize(target)==0: os.remove(target)
        except OSError: pass
        return None

def maybe_backup_database():
    global _backup_last_check
    now=time.time()
    if now-_backup_last_check < 3600:
        return
    _backup_last_check=now
    backups=[os.path.join(BACKUP_DIR,x) for x in os.listdir(BACKUP_DIR)
             if x.startswith("kk_") and x.endswith(".db")]
    newest=max([os.path.getmtime(x) for x in backups], default=0)
    if now-newest >= 24*3600:
        backup_database()





def admin_csrf_token():
    token=session.get("_admin_csrf")
    if not token:
        token=secrets.token_urlsafe(32)
        session["_admin_csrf"]=token
    return token

def valid_admin_csrf():
    # Accept the token from the standard header, normal form submissions,
    # or JSON bodies. The JSON fallback makes AJAX actions resilient to
    # environments/proxies that strip custom headers.
    supplied=request.headers.get("X-CSRF-Token") or request.form.get("_csrf")
    if not supplied and request.is_json:
        try:
            payload=request.get_json(silent=True) or {}
            supplied=payload.get("_csrf")
        except Exception:
            supplied=None
    stored=session.get("_admin_csrf","")
    return bool(supplied and stored and hmac.compare_digest(str(supplied),str(stored)))

def get_admin_account():
    c=conn()
    row=c.execute("SELECT id,username,password_hash,session_version FROM admin_account WHERE id=1").fetchone()
    if row is None and ADMIN_PASSWORD_HASH:
        c.execute("INSERT INTO admin_account(id,username,password_hash,session_version) VALUES(1,?,?,1)",
                  (ADMIN_USERNAME,ADMIN_PASSWORD_HASH))
        c.commit()
        row=c.execute("SELECT id,username,password_hash,session_version FROM admin_account WHERE id=1").fetchone()
    c.close()
    return row

def admin_password_strong(password):
    return (
        len(password)>=12
        and any(ch.isalpha() for ch in password)
        and any(ch.isdigit() for ch in password)
        and any(not ch.isalnum() for ch in password)
    )

@app.context_processor
def admin_security_context():
    return {"admin_csrf_token": admin_csrf_token() if session.get("admin") else ""}

@app.before_request
def global_security_guard():
    maybe_backup_database()
    if FORCE_HTTPS and not request.is_secure:
        return redirect(request.url.replace("http://","https://",1), code=308)
    if request.path.startswith("/api/orders") and request.method=="POST":
        if not online_order_enabled():
            return jsonify(ok=False,message="K&K hiện đang tạm ngưng nhận đặt món trực tuyến. Quý khách vui lòng liên hệ K&K để được hỗ trợ."),403
        if not consume_rate_limit("order"):
            return jsonify(ok=False,message="Bạn thao tác quá nhanh. Vui lòng thử lại sau ít phút."),429
    # Voucher chỉ phục vụ luồng đặt hàng, nên cũng phải tắt khi đặt hàng online đang tắt.
    if request.path == "/api/voucher" and request.method == "POST" and not online_order_enabled():
        return jsonify(ok=False,message="K&K hiện đang tạm ngưng đặt hàng trực tuyến."),403
    # Tính năng trực tuyến dùng chung một công tắc: Đặt hàng + Đặt bàn + Theo dõi đơn.
    if request.path in ("/book", "/track") and not online_order_enabled():
        return redirect("/")
    if request.path in ("/api/bookings", "/api/track") and not online_order_enabled():
        return jsonify(ok=False,message="K&K hiện đang tạm ngưng các tính năng trực tuyến. Quý khách vui lòng liên hệ K&K để được hỗ trợ."),403
    if request.path.startswith("/api/bookings") and request.method=="POST":
        if not consume_rate_limit("booking"):
            return jsonify(ok=False,message="Bạn gửi quá nhiều yêu cầu đặt bàn. Vui lòng thử lại sau ít phút."),429
    return None

@app.before_request
def admin_security_guard():
    if not request.path.startswith("/admin"):
        return None

    # Login is the only unauthenticated POST endpoint.
    if request.path == "/admin/login":
        return None

    if not session.get("admin"):
        if request.path == "/admin/logout":
            return redirect("/admin")
        return None

    # 30-minute idle timeout.
    now=datetime.now().timestamp()
    last=float(session.get("admin_last_activity",now))
    if now-last > ADMIN_IDLE_TIMEOUT:
        session.clear()
        if is_ajax_request():
            return jsonify(ok=False,message="Phiên quản trị đã hết hạn do không hoạt động."), 401
        return redirect("/admin")

    account=get_admin_account()
    if not account or int(session.get("admin_session_version",0)) != int(account["session_version"]):
        session.clear()
        if is_ajax_request():
            return jsonify(ok=False,message="Phiên quản trị đã bị vô hiệu hóa. Vui lòng đăng nhập lại."), 401
        return redirect("/admin")

    # All Admin POST actions except login require CSRF.
    if request.method=="POST" and not valid_admin_csrf():
        return jsonify(ok=False,message="Yêu cầu bảo mật không hợp lệ (CSRF). Hãy tải lại trang Admin."),403

    session["admin_last_activity"]=now
    return None

@app.after_request
def security_headers(response):
    response.headers.setdefault("X-Content-Type-Options","nosniff")
    response.headers.setdefault("X-Frame-Options","SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy","strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy","camera=(), microphone=(), geolocation=()")
    if HSTS_ENABLED and request.is_secure:
        response.headers.setdefault("Strict-Transport-Security","max-age=31536000; includeSubDomains")
    if CSP_ENABLED:
        response.headers.setdefault(
            "Content-Security-Policy",
            "default-src 'self'; img-src 'self' data: blob:; style-src 'self' 'unsafe-inline'; "
            "script-src 'self' 'unsafe-inline'; connect-src 'self'; font-src 'self' data:; "
            "frame-ancestors 'self'; base-uri 'self'; form-action 'self'"
        )
    if request.path.startswith("/admin"):
        response.headers.setdefault("Cache-Control","no-store")
        response.headers.setdefault("X-Robots-Tag","noindex, nofollow")
    return response

def is_ajax_request():
    return request.headers.get("X-Requested-With") == "XMLHttpRequest" or "application/json" in (request.headers.get("Accept") or "")

def admin_result(message="Đã cập nhật.", refresh=True, status=200, **extra):
    if is_ajax_request():
        return jsonify(ok=(status < 400), message=message, refresh=refresh, **extra), status
    return None

def conn():
    c=sqlite3.connect(DB, timeout=10, isolation_level="DEFERRED")
    c.row_factory=sqlite3.Row
    c.execute("PRAGMA busy_timeout=10000")
    c.execute("PRAGMA foreign_keys=ON")
    c.execute("PRAGMA journal_mode=WAL")
    c.execute("PRAGMA synchronous=NORMAL")
    return c

def init_db():
    os.makedirs(os.path.dirname(DB),exist_ok=True); os.makedirs(UP,exist_ok=True)
    c=conn()
    c.execute("""CREATE TABLE IF NOT EXISTS admin_account(
        id INTEGER PRIMARY KEY CHECK(id=1),
        username TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        session_version INTEGER NOT NULL DEFAULT 1,
        updated_at TEXT DEFAULT CURRENT_TIMESTAMP
    )""")
    if c.execute("SELECT COUNT(*) n FROM admin_account").fetchone()["n"]==0 and ADMIN_PASSWORD_HASH:
        c.execute("INSERT INTO admin_account(id,username,password_hash,session_version) VALUES(1,?,?,1)",
                  (ADMIN_USERNAME,ADMIN_PASSWORD_HASH))
    c.execute("""CREATE TABLE IF NOT EXISTS menu_items(
        id TEXT PRIMARY KEY,name TEXT,category TEXT,price INTEGER,description TEXT,
        image TEXT,badge TEXT,active INTEGER DEFAULT 1,
        featured INTEGER DEFAULT 0,sort_order INTEGER DEFAULT 0)""")
    # V2.2 migration: keep existing databases working while adding featured/order.
    menu_cols={r["name"] for r in c.execute("PRAGMA table_info(menu_items)").fetchall()}
    if "featured" not in menu_cols:
        c.execute("ALTER TABLE menu_items ADD COLUMN featured INTEGER DEFAULT 0")
    if "sort_order" not in menu_cols:
        c.execute("ALTER TABLE menu_items ADD COLUMN sort_order INTEGER DEFAULT 0")
    c.execute("""CREATE TABLE IF NOT EXISTS categories(
        id TEXT PRIMARY KEY,name TEXT NOT NULL,sort_order INTEGER DEFAULT 0,active INTEGER DEFAULT 1)""")
    c.execute("""CREATE TABLE IF NOT EXISTS orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,order_code TEXT UNIQUE,customer_name TEXT,phone TEXT,
        address TEXT,note TEXT,items TEXT,subtotal INTEGER,discount INTEGER,total INTEGER,
        payment_method TEXT,payment_status TEXT,status TEXT,created_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS vouchers(
        code TEXT PRIMARY KEY,type TEXT,value INTEGER,min_order INTEGER DEFAULT 0,
        active INTEGER DEFAULT 1,start_date TEXT,end_date TEXT)""")
    voucher_cols={r["name"] for r in c.execute("PRAGMA table_info(vouchers)").fetchall()}
    if "start_date" not in voucher_cols:
        c.execute("ALTER TABLE vouchers ADD COLUMN start_date TEXT")
    if "end_date" not in voucher_cols:
        c.execute("ALTER TABLE vouchers ADD COLUMN end_date TEXT")
    c.execute("""CREATE TABLE IF NOT EXISTS bookings(
        id INTEGER PRIMARY KEY AUTOINCREMENT,name TEXT,phone TEXT,booking_date TEXT,
        booking_time TEXT,guests INTEGER,note TEXT,status TEXT,created_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS settings(k TEXT PRIMARY KEY,v TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS rate_limit_hits(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scope TEXT NOT NULL,
        key_hash TEXT NOT NULL,
        created_at INTEGER NOT NULL
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS audit_logs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT NOT NULL,
        username TEXT,
        ip TEXT,
        action TEXT NOT NULL,
        entity_type TEXT,
        entity_id TEXT,
        details TEXT
    )""")
    order_cols={r["name"] for r in c.execute("PRAGMA table_info(orders)").fetchall()}
    if "idempotency_key" not in order_cols:
        c.execute("ALTER TABLE orders ADD COLUMN idempotency_key TEXT")
    # Categories must exist before menu rows are inserted because menu_items
    # is migrated to enforce a foreign key to categories below.
    if c.execute("SELECT COUNT(*) n FROM categories").fetchone()["n"]==0:
        for i,(cid,cname) in enumerate(DEFAULT_CATEGORIES):
            c.execute("INSERT INTO categories VALUES(?,?,?,1)",(cid,cname,i))
    # Migrate legacy menu_items to a real foreign key once categories exist.
    fk_menu=c.execute("PRAGMA foreign_key_list(menu_items)").fetchall()
    orphan_count=c.execute("""SELECT COUNT(*) n FROM menu_items m
                              LEFT JOIN categories cat ON cat.id=m.category
                              WHERE cat.id IS NULL""").fetchone()["n"]
    if not fk_menu:
        if orphan_count:
            if not c.execute("SELECT 1 FROM categories WHERE id='__other__'").fetchone():
                n=c.execute("SELECT COALESCE(MAX(sort_order),-1)+1 n FROM categories").fetchone()["n"]
                c.execute("INSERT INTO categories(id,name,sort_order,active) VALUES('__other__','Chưa phân loại / Khác',?,0)",(n,))
            c.execute("UPDATE menu_items SET category='__other__' WHERE category NOT IN (SELECT id FROM categories)")
        c.execute("ALTER TABLE menu_items RENAME TO menu_items_legacy")
        c.execute("""CREATE TABLE menu_items(
            id TEXT PRIMARY KEY,name TEXT,category TEXT,price INTEGER,description TEXT,
            image TEXT,badge TEXT,active INTEGER DEFAULT 1,
            featured INTEGER DEFAULT 0,sort_order INTEGER DEFAULT 0,
            FOREIGN KEY(category) REFERENCES categories(id) ON UPDATE CASCADE ON DELETE RESTRICT
        )""")
        c.execute("""INSERT INTO menu_items(id,name,category,price,description,image,badge,active,featured,sort_order)
                     SELECT id,name,category,price,description,image,badge,active,featured,sort_order
                     FROM menu_items_legacy""")
        c.execute("DROP TABLE menu_items_legacy")
    if c.execute("SELECT COUNT(*) n FROM menu_items").fetchone()["n"]==0:
        for x in SEED_MENU:
            c.execute("""INSERT INTO menu_items
                (id,name,category,price,description,image,badge,active,featured,sort_order)
                VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (x["id"],x["name"],x["category"],x["price"],x["desc"],
                 f"/static/dishes/{x['id']}.svg",x.get("badge",""),1,0,0))
    for k,v in DEFAULT_SETTINGS.items():
        c.execute("INSERT OR IGNORE INTO settings VALUES(?,?)",(k,v))
    # Give legacy rows a stable order only when the column is still at its
    # untouched default (0 for every row). Never reset an order already chosen
    # by an administrator on every application restart.
    order_stats=c.execute("SELECT COUNT(*) n, COALESCE(MIN(sort_order),0) mn, COALESCE(MAX(sort_order),0) mx FROM menu_items").fetchone()
    if order_stats["n"] > 1 and order_stats["mx"] == 0:
        c.execute("UPDATE menu_items SET sort_order=rowid-1")
    c.execute("""INSERT OR IGNORE INTO vouchers
        (code,type,value,min_order,active,start_date,end_date)
        VALUES ('WELCOME10','percent',10,3000,1,NULL,NULL)""")
    c.execute("""INSERT OR IGNORE INTO vouchers
        (code,type,value,min_order,active,start_date,end_date)
        VALUES ('KK50','fixed',500,5000,1,NULL,NULL)""")
    c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_orders_idempotency ON orders(idempotency_key) WHERE idempotency_key IS NOT NULL")
    c.execute("CREATE INDEX IF NOT EXISTS idx_orders_status_created ON orders(status,created_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_orders_phone_code ON orders(phone,order_code)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_bookings_date_status ON bookings(booking_date,status)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_menu_category_active_order ON menu_items(category,active,sort_order)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_categories_active_order ON categories(active,sort_order)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_rate_limit_lookup ON rate_limit_hits(scope,key_hash,created_at)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_audit_logs_created ON audit_logs(created_at)")
    c.commit(); c.close()

# Initialize/migrate the local SQLite schema when the module is imported.
# Always keep a pre-migration snapshot when a database already exists.
if os.path.isfile(DB) and os.getenv("DISABLE_AUTO_BACKUP","0")!="1":
    backup_database()
init_db()
if os.getenv("DISABLE_AUTO_BACKUP","0")!="1":
    backup_database()

def settings():
    c=conn(); rows=c.execute("SELECT k,v FROM settings").fetchall(); c.close()
    data={r["k"]:r["v"] for r in rows}
    for k,v in DEFAULT_SETTINGS.items():
        data.setdefault(k,v)
    return data

def online_order_enabled():
    return str(settings().get("online_order_enabled", "0")).strip().lower() in ("1", "true", "yes", "on")

def categories():
    c=conn(); rows=c.execute("SELECT * FROM categories WHERE active=1 ORDER BY sort_order,id").fetchall(); c.close()
    return [dict(r) for r in rows]

def category_choices():
    return [(r["id"],r["name"]) for r in categories()]

def menu():
    c=conn(); rows=c.execute("""SELECT m.* FROM menu_items m
        JOIN categories c ON c.id=m.category
        WHERE m.active=1 AND c.active=1
        ORDER BY c.sort_order,m.sort_order,m.name""").fetchall(); c.close()
    return [dict(r) for r in rows]

def featured_menu():
    c=conn(); rows=c.execute("""SELECT m.* FROM menu_items m
        JOIN categories c ON c.id=m.category
        WHERE m.active=1 AND m.featured=1 AND c.active=1
        ORDER BY m.sort_order,m.name LIMIT 8""").fetchall(); c.close()
    return [dict(r) for r in rows]

@app.context_processor
def inject():
    return {"s":settings(),"restaurant_name":"Phở & Bánh Mì K&K","online_order_enabled":online_order_enabled()}

@app.route("/")
def home():
    featured=featured_menu()
    return render_template("index.html",menu=featured or menu()[:8],categories=category_choices())

@app.route("/menu")
def menu_page(): return render_template("menu.html",menu=menu(),categories=[("all","Tất cả")]+category_choices())

@app.route("/checkout")
def checkout(): return render_template("checkout.html")

@app.route("/track")
def track(): return render_template("track.html")

@app.route("/book")
def book(): return render_template("book.html")

def calculate_voucher(code, subtotal):
    """Return (ok, discount, message) using only server-side voucher data."""
    code=(code or "").strip().upper()
    try:
        subtotal=max(0,int(subtotal))
    except (TypeError,ValueError):
        return False,0,"Giá trị đơn hàng không hợp lệ."
    if not code:
        return False,0,"Vui lòng nhập mã voucher."
    c=conn()
    row=c.execute("SELECT * FROM vouchers WHERE code=? AND active=1",(code,)).fetchone()
    c.close()
    today=datetime.now().date().isoformat()
    if not row or (row["start_date"] and today < row["start_date"]) or (row["end_date"] and today > row["end_date"]):
        return False,0,"Voucher không hợp lệ hoặc chưa/đã hết thời gian áp dụng."
    min_order=max(0,int(row["min_order"] or 0))
    if subtotal < min_order:
        return False,0,f"Đơn tối thiểu {money(min_order)}."
    value=max(0,int(row["value"] or 0))
    if row["type"]=="percent":
        discount=round(subtotal*value/100)
    else:
        discount=value
    discount=min(max(0,discount),subtotal)
    return True,discount,f"Áp dụng {code} thành công."

@app.post("/api/voucher")
def voucher():
    d=request.json or {}
    ok,discount,message=calculate_voucher(d.get("code"),d.get("subtotal",0))
    return jsonify(ok=ok,discount=discount if ok else 0,message=message), (200 if ok else 400)

@app.post("/api/orders")
def create_order():
    d=request.json or {}
    required=["customer_name","phone","address","items"]
    if not all(str(d.get(x) or "").strip() for x in required):
        return jsonify(ok=False,message="Vui lòng điền đủ thông tin."),400
    if d.get("payment_method") not in (None,"","cod"):
        return jsonify(ok=False,message="Phương thức thanh toán không hợp lệ."),400
    items=d.get("items")
    if not isinstance(items,list) or not items:
        return jsonify(ok=False,message="Giỏ hàng đang trống."),400

    idempotency_key=str(d.get("idempotency_key") or "").strip()
    if not idempotency_key:
        return jsonify(ok=False,message="Thiếu mã chống tạo đơn trùng. Vui lòng tải lại trang và thử lại."),400
    if len(idempotency_key)>100:
        return jsonify(ok=False,message="Mã yêu cầu không hợp lệ."),400

    requested={}
    for raw in items:
        if not isinstance(raw,dict):
            return jsonify(ok=False,message="Dữ liệu giỏ hàng không hợp lệ."),400
        item_id=str(raw.get("id") or "").strip()
        try:
            qty=int(raw.get("qty") or 0)
        except (TypeError,ValueError):
            qty=0
        if not item_id or qty<1 or qty>99:
            return jsonify(ok=False,message="Số lượng món không hợp lệ."),400
        requested[item_id]=requested.get(item_id,0)+qty
        if requested[item_id]>99:
            return jsonify(ok=False,message="Số lượng một món không được vượt quá 99."),400

    c=conn()
    try:
        # Idempotency: a retried request returns the already-created order.
        existing=c.execute("""SELECT order_code
                             FROM orders WHERE idempotency_key=?""",(idempotency_key,)).fetchone()
        if existing:
            return jsonify(ok=True,order_code=existing["order_code"],idempotent=True)

        placeholders=",".join("?" for _ in requested)
        rows=c.execute(
            f"""SELECT m.id,m.name,m.price,m.active,cat.active AS category_active
                FROM menu_items m
                JOIN categories cat ON cat.id=m.category
                WHERE m.id IN ({placeholders})""",
            tuple(requested.keys())
        ).fetchall() if requested else []
        by_id={r["id"]:r for r in rows}
        if len(by_id)!=len(requested) or any(
            not by_id[i]["active"] or not by_id[i]["category_active"] for i in requested
        ):
            return jsonify(ok=False,message="Một hoặc nhiều món trong giỏ hàng không còn được bán hoặc danh mục đã ẩn. Vui lòng cập nhật lại giỏ hàng."),409

        normalized=[]
        sub=0
        for item_id,qty in requested.items():
            row=by_id[item_id]
            price=max(0,int(row["price"] or 0))
            normalized.append({"id":row["id"],"name":row["name"],"price":price,"qty":qty})
            sub += price*qty

        voucher_code=(d.get("voucher") or "").strip().upper()
        disc=0
        if voucher_code:
            vok,disc,vmsg=calculate_voucher(voucher_code,sub)
            if not vok:
                return jsonify(ok=False,message=vmsg),400
        total=max(0,sub-disc)
        code="KK"+datetime.now().strftime("%y%m%d")+"-"+secrets.token_hex(4).upper()

        try:
            cur=c.execute("""INSERT INTO orders
                (order_code,customer_name,phone,address,note,items,subtotal,discount,total,
                 payment_method,payment_status,status,created_at,idempotency_key)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (code,str(d["customer_name"]).strip(),str(d["phone"]).strip(),str(d["address"]).strip(),
                 str(d.get("note") or "").strip(),json.dumps(normalized,ensure_ascii=False),
                 sub,disc,total,"cod","COD",
                 "Mới",datetime.now().strftime("%Y-%m-%d %H:%M:%S"),idempotency_key))
            c.commit()
        except sqlite3.IntegrityError:
            c.rollback()
            existing=c.execute("""SELECT order_code
                                 FROM orders WHERE idempotency_key=?""",(idempotency_key,)).fetchone()
            if existing:
                return jsonify(ok=True,order_code=existing["order_code"],idempotent=True)
            raise
        return jsonify(ok=True,order_code=code)
    except Exception:
        c.rollback()
        logger.exception("create_order failed")
        return jsonify(ok=False,message="Không thể tạo đơn hàng. Vui lòng thử lại."),500
    finally:
        c.close()

@app.post("/api/bookings")
def create_booking():
    d=request.json or {}
    required=["name","phone","booking_date","booking_time","guests"]
    if not all(str(d.get(x) or "").strip() for x in required):
        return jsonify(ok=False,message="Vui lòng điền đủ thông tin."),400
    try:
        booking_date=datetime.strptime(str(d["booking_date"]),"%Y-%m-%d").date()
        booking_time=datetime.strptime(str(d["booking_time"]),"%H:%M").time()
        guests=int(d["guests"])
    except (TypeError,ValueError):
        return jsonify(ok=False,message="Ngày, giờ hoặc số khách không hợp lệ."),400
    if booking_date < datetime.now().date():
        return jsonify(ok=False,message="Ngày đặt bàn không được ở trong quá khứ."),400
    if not 1 <= guests <= 50:
        return jsonify(ok=False,message="Số khách phải từ 1 đến 50."),400
    c=conn()
    cur=c.execute("""INSERT INTO bookings
        (name,phone,booking_date,booking_time,guests,note,status,created_at)
        VALUES(?,?,?,?,?,?,?,?)""",
        (str(d["name"]).strip(),str(d["phone"]).strip(),booking_date.isoformat(),booking_time.strftime("%H:%M"),
         guests,str(d.get("note") or "").strip(),"Mới",datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    bid=cur.lastrowid; c.commit(); c.close()
    logger.info("New booking %s from %s", bid, client_ip())
    return jsonify(ok=True,booking_id=bid)

@app.post("/api/track")
def api_track():
    d=request.json or {}; code=(d.get("order_code") or "").strip().upper(); phone=(d.get("phone") or "").strip()
    c=conn(); row=c.execute("""SELECT order_code,status,payment_status,customer_name,total,created_at
        FROM orders WHERE order_code=? AND phone=?""",(code,phone)).fetchone(); c.close()
    if not row: return jsonify(ok=False,message="Không tìm thấy đơn hàng. Kiểm tra mã đơn và số điện thoại.")
    return jsonify(ok=True,order=dict(row))

def admin_context():
    c=conn()
    orders=c.execute("SELECT * FROM orders ORDER BY id DESC").fetchall()
    bookings=c.execute("SELECT * FROM bookings ORDER BY id DESC").fetchall()
    vouchers=[dict(r) for r in c.execute("SELECT * FROM vouchers ORDER BY code").fetchall()]
    items=[dict(r) for r in c.execute("SELECT * FROM menu_items ORDER BY sort_order,category,name").fetchall()]
    cats=[dict(r) for r in c.execute("SELECT * FROM categories ORDER BY sort_order,id").fetchall()]
    category_options=[("all","Tất cả")]+[(r["id"],r["name"]) for r in cats]
    category_names={r["id"]:r["name"] for r in cats}
    category_groups=[]; grouped_ids=set()
    for cat in cats:
        cat_id=cat["id"]; group_items=[x for x in items if x.get("category")==cat_id]; grouped_ids.add(cat_id)
        category_groups.append({"id":cat_id,"name":cat["name"],"active":cat.get("active",1),"items":group_items})
    orphan_items=[x for x in items if x.get("category") not in grouped_ids]
    if orphan_items:
        category_groups.append({"id":"__other__","name":"Chưa phân loại / Khác","active":1,"items":orphan_items})
    featured_items=[dict(r) for r in c.execute("""SELECT m.* FROM menu_items m
        JOIN categories c ON c.id=m.category
        WHERE m.active=1 AND c.active=1
        ORDER BY m.sort_order,m.name""").fetchall()]
    site_settings={r["k"]:r["v"] for r in c.execute("SELECT k,v FROM settings").fetchall()}
    for k,v in DEFAULT_SETTINGS.items():
        site_settings.setdefault(k,v)
    c.close()
    account=get_admin_account()
    return dict(orders=orders, bookings=bookings, vouchers=vouchers, items=items, cats=cats,
                categories=category_options, category_names=category_names,
                category_groups=category_groups, featured_items=featured_items,
                site_settings=site_settings, admin_account=account,
                admin_csrf_token=admin_csrf_token())

@app.route("/admin")
def admin():
    if not session.get("admin"): return render_template("admin_login.html")
    ctx=admin_context()
    if request.args.get("partial") == "1":
        return render_template("admin_body.html", **ctx)
    return render_template("admin.html", **ctx)

@app.post("/admin/login")
def login():
    username=(request.form.get("username") or "").strip()
    password=request.form.get("password") or ""
    if not consume_rate_limit("login", username.lower()[:80]):
        return render_template("admin_login.html",error="Đăng nhập tạm thời bị giới hạn. Vui lòng thử lại sau 15 phút."),429

    account=get_admin_account()
    if account and username==account["username"] and check_password_hash(account["password_hash"],password):
        session.clear()
        session["admin"]=True
        session["admin_username"]=account["username"]
        session["admin_session_version"]=int(account["session_version"])
        session["admin_last_activity"]=datetime.now().timestamp()
        session.permanent=True
        audit_log("LOGIN_SUCCESS","admin",account["id"],{"username":account["username"]})
        return redirect("/admin")

    return render_template("admin_login.html",error="Sai tài khoản hoặc mật khẩu.")

@app.get("/admin/logout")
def logout():
    session.clear()
    return redirect("/admin")

@app.post("/admin/security/update")
def admin_security_update():
    account=get_admin_account()
    if not account:
        return jsonify(ok=False,message="Chưa thiết lập tài khoản quản trị."),500

    username=(request.form.get("username") or "").strip()
    current_password=request.form.get("current_password") or ""
    new_password=request.form.get("new_password") or ""
    confirm_password=request.form.get("confirm_password") or ""

    if not (3 <= len(username) <= 50):
        return jsonify(ok=False,message="Tên đăng nhập phải từ 3 đến 50 ký tự."),400
    if not check_password_hash(account["password_hash"],current_password):
        return jsonify(ok=False,message="Mật khẩu hiện tại không đúng."),400
    if not admin_password_strong(new_password):
        return jsonify(ok=False,message="Mật khẩu mới phải có ít nhất 12 ký tự, gồm chữ, số và ký tự đặc biệt."),400
    if new_password != confirm_password:
        return jsonify(ok=False,message="Mật khẩu xác nhận không khớp."),400

    new_hash=generate_password_hash(new_password)
    new_version=int(account["session_version"])+1
    c=conn()
    c.execute("UPDATE admin_account SET username=?,password_hash=?,session_version=?,updated_at=CURRENT_TIMESTAMP WHERE id=1",
              (username,new_hash,new_version))
    c.commit(); c.close()

    session["admin"]=True
    session["admin_username"]=username
    session["admin_session_version"]=new_version
    session["admin_last_activity"]=datetime.now().timestamp()
    session.permanent=True
    audit_log("CHANGE_CREDENTIALS","admin",1,{"username":username})
    return jsonify(ok=True,message="Đã đổi tài khoản và mật khẩu quản trị.")

@app.post("/admin/security/logout-other-sessions")
def admin_logout_other_sessions():
    account=get_admin_account()
    if not account:
        return jsonify(ok=False,message="Chưa thiết lập tài khoản quản trị."),500
    new_version=int(account["session_version"])+1
    c=conn()
    c.execute("UPDATE admin_account SET session_version=?,updated_at=CURRENT_TIMESTAMP WHERE id=1",(new_version,))
    c.commit(); c.close()
    session["admin_session_version"]=new_version
    session["admin_last_activity"]=datetime.now().timestamp()
    audit_log("REVOKE_SESSIONS","admin",1,{})
    return jsonify(ok=True,message="Đã vô hiệu hóa các phiên quản trị khác.")

@app.post("/admin/order/<int:oid>/status")
def order_status(oid):
    if not session.get("admin"): return jsonify(ok=False),401
    data=request.json or {}; st=data.get("status","Mới")
    if st not in ORDER_STATUSES:
        return jsonify(ok=False,message="Trạng thái đơn hàng không hợp lệ."),400
    c=conn(); cur=c.execute("UPDATE orders SET status=? WHERE id=?",(st,oid))
    c.commit(); c.close()
    if cur.rowcount == 1:
        audit_log("UPDATE_STATUS","order",oid,{"status":st})
    if cur.rowcount == 0:
        return jsonify(ok=False,message="Không tìm thấy đơn hàng."),404
    return jsonify(ok=True,message="Đã cập nhật trạng thái đơn hàng.")

@app.post("/admin/booking/<int:bid>/status")
def booking_status(bid):
    if not session.get("admin"): return jsonify(ok=False),401
    data=request.json or {}; st=data.get("status","Mới")
    if st not in BOOKING_STATUSES:
        return jsonify(ok=False,message="Trạng thái đặt bàn không hợp lệ."),400
    c=conn(); cur=c.execute("UPDATE bookings SET status=? WHERE id=?",(st,bid))
    c.commit(); c.close()
    if cur.rowcount == 1:
        audit_log("UPDATE_STATUS","booking",bid,{"status":st})
    if cur.rowcount == 0:
        return jsonify(ok=False,message="Không tìm thấy đặt bàn."),404
    return jsonify(ok=True,message="Đã cập nhật trạng thái đặt bàn.")

def save_menu_item(item_id):
    if not re.fullmatch(r"[A-Za-z0-9_-]{1,80}", str(item_id or "")):
        raise ValueError("Mã món không hợp lệ.")
    name=(request.form.get("name") or "").strip()
    category=(request.form.get("category") or "").strip()
    description=(request.form.get("description") or "").strip()
    badge=(request.form.get("badge") or "").strip()
    if not name:
        raise ValueError("Vui lòng nhập tên món.")
    try:
        price=parse_money_cents(request.form.get("price"))
    except ValueError:
        raise ValueError("Giá món phải là số tiền hợp lệ, ví dụ 9.90.")

    c=conn()
    cat_exists=c.execute("SELECT 1 FROM categories WHERE id=?",(category,)).fetchone()
    if not cat_exists:
        c.close()
        raise ValueError("Danh mục món ăn không hợp lệ.")

    old_row=c.execute("SELECT image,sort_order,featured FROM menu_items WHERE id=?",(item_id,)).fetchone()
    f=request.files.get("image")
    submitted_old=(request.form.get("old_image") or "").strip()
    if submitted_old and not (submitted_old.startswith("/static/dishes/") or submitted_old.startswith("/uploads/")):
        submitted_old=""
    image=submitted_old or (old_row["image"] if old_row else f"/static/dishes/{item_id}.svg")
    saved_file=None
    if f and f.filename:
        original=secure_filename(f.filename)
        ext=original.rsplit(".",1)[-1].lower() if "." in original else ""
        if ext not in IMAGE_EXTENSIONS:
            c.close()
            raise ValueError("Ảnh món phải có định dạng JPG, JPEG, PNG hoặc WEBP.")
        filename=f"{item_id}_{secrets.token_hex(6)}.{ext}"
        saved_file=os.path.join(UP,filename)
        save_validated_image(f, saved_file, ext)
        image="/uploads/"+filename

    try:
        sort_order=int(request.form.get("sort_order") or (old_row["sort_order"] if old_row else 0))
    except (TypeError,ValueError):
        sort_order=int(old_row["sort_order"] if old_row else 0)

    active=1 if request.form.get("active") else 0
    featured=1 if request.form.get("featured") and active else 0
    try:
        c.execute("""INSERT OR REPLACE INTO menu_items
            (id,name,category,price,description,image,badge,active,featured,sort_order)
            VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (item_id,name,category,price,description,image,badge,active,featured,sort_order))
        c.commit()
    except Exception:
        c.rollback(); c.close()
        if saved_file:
            try: os.remove(saved_file)
            except OSError: pass
        raise
    c.close()

    if saved_file and old_row and old_row["image"].startswith("/uploads/") and old_row["image"]!=image:
        try: os.remove(os.path.join(UP,os.path.basename(old_row["image"])))
        except OSError: pass


IMPORT_COLUMNS=["Category","Name","Price_USD","Description","Badge","Available","Featured","Image"]


def _clean_import_value(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_import_bool(value, default=True):
    raw=_clean_import_value(value).strip().lower()
    if not raw:
        return 1 if default else 0
    if raw in {"yes","y","true","1","on","active","available","đang bán","có"}:
        return 1
    if raw in {"no","n","false","0","off","inactive","unavailable","tạm ẩn","không"}:
        return 0
    raise ValueError("chỉ nhận Yes/No")


def _resolve_import_category(value, category_rows):
    raw=_clean_import_value(value)
    if not raw:
        raise ValueError("thiếu Category")
    candidates=[raw]
    if "|" in raw:
        candidates.extend([x.strip() for x in raw.split("|") if x.strip()])
    lookup={}
    for row in category_rows:
        cid=str(row["id"]).strip()
        name=str(row["name"]).strip()
        lookup[cid.lower()]=cid
        lookup[name.lower()]=cid
        # Allows bilingual cells such as: Phở | Noodle Soup.
        for part in name.split("|"):
            lookup[part.strip().lower()]=cid
    for candidate in candidates:
        cid=lookup.get(candidate.lower())
        if cid:
            return cid
    raise ValueError(f"danh mục không tồn tại: {raw}")


def _resolve_import_image(value):
    raw=_clean_import_value(value)
    if not raw:
        return ""
    if raw.startswith(("/uploads/","/static/dishes/")):
        public=raw
        disk=os.path.join(BASE, raw.lstrip("/"))
        if os.path.isfile(disk):
            return public
        raise ValueError(f"ảnh không tồn tại: {raw}")
    safe=secure_filename(raw)
    if not safe or safe != os.path.basename(raw):
        raise ValueError("Image chỉ được chứa tên file, không được chứa đường dẫn")
    for directory, prefix in ((UP,"/uploads/"),(os.path.join(BASE,"static","dishes"),"/static/dishes/")):
        path=os.path.join(directory,safe)
        if os.path.isfile(path):
            return prefix+safe
    raise ValueError(f"không tìm thấy ảnh trong uploads hoặc static/dishes: {safe}")


def _cleanup_import_staging(max_age=1800):
    now=time.time()
    try:
        for name in os.listdir(IMPORT_DIR):
            path=os.path.join(IMPORT_DIR,name)
            if os.path.isfile(path) and now-os.path.getmtime(path)>max_age:
                try: os.remove(path)
                except OSError: pass
    except OSError:
        pass


def _validate_import_workbook(file_obj):
    if load_workbook is None:
        raise ValueError("Chưa cài openpyxl. Hãy chạy: py -m pip install openpyxl")
    filename=secure_filename(file_obj.filename or "")
    if not filename.lower().endswith(".xlsx"):
        raise ValueError("Chỉ hỗ trợ file Excel .xlsx")
    try:
        file_obj.stream.seek(0)
        wb=load_workbook(file_obj.stream, read_only=True, data_only=True)
        ws=wb["Menu"] if "Menu" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows=ws.iter_rows(values_only=True)
        header=next(rows,None)
        normalized=[_clean_import_value(v) for v in (header or ())]
        if normalized[:len(IMPORT_COLUMNS)] != IMPORT_COLUMNS or len(normalized)<len(IMPORT_COLUMNS):
            raise ValueError("Header Excel không đúng. Hãy dùng đúng file mẫu K&K và không đổi tên 8 cột.")

        c=conn()
        category_rows=c.execute("SELECT id,name,active FROM categories ORDER BY sort_order,id").fetchall()
        existing=c.execute("SELECT id,name,category,sort_order,image,featured FROM menu_items").fetchall()
        c.close()
        existing_map={}
        existing_dupes=set()
        for row in existing:
            key=(str(row["category"]).lower(), str(row["name"]).strip().casefold())
            if key in existing_map:
                existing_dupes.add(key)
            else:
                existing_map[key]=row

        parsed=[]; errors=[]; warnings=[]; seen=set()
        for excel_row_num, values in enumerate(rows, start=2):
            vals=list(values or ())
            if not any(_clean_import_value(v) for v in vals):
                continue
            if excel_row_num>MAX_IMPORT_ROWS+1:
                errors.append({"row":excel_row_num,"message":f"Vượt giới hạn {MAX_IMPORT_ROWS} dòng dữ liệu."})
                break
            vals=(vals+[None]*len(IMPORT_COLUMNS))[:len(IMPORT_COLUMNS)]
            raw={IMPORT_COLUMNS[i]:vals[i] for i in range(len(IMPORT_COLUMNS))}
            row_errors=[]
            name=_clean_import_value(raw["Name"])
            if not name: row_errors.append("thiếu Name")
            try: category=_resolve_import_category(raw["Category"],category_rows)
            except ValueError as exc: category=""; row_errors.append(str(exc))
            price_raw=_clean_import_value(raw["Price_USD"])
            try:
                if not price_raw: raise ValueError("thiếu Price_USD")
                price=parse_money_cents(price_raw)
                if price<=0: raise ValueError("Price_USD phải lớn hơn 0")
            except ValueError as exc: price=0; row_errors.append(str(exc))
            try: active=_parse_import_bool(raw["Available"],True)
            except ValueError as exc: active=1; row_errors.append(f"Available {exc}")
            try: featured=_parse_import_bool(raw["Featured"],False)
            except ValueError as exc: featured=0; row_errors.append(f"Featured {exc}")
            image=""
            if _clean_import_value(raw["Image"]):
                try: image=_resolve_import_image(raw["Image"])
                except ValueError as exc: row_errors.append(str(exc))
            key=(category.lower(),name.casefold()) if category and name else None
            if key and key in seen: row_errors.append("trùng Category + Name trong cùng file")
            if key: seen.add(key)
            existing_row=existing_map.get(key) if key else None
            if key and key in existing_dupes:
                row_errors.append("database đang có nhiều món trùng Category + Name; hãy xử lý trùng trước khi import")
            cat_active=next((int(r["active"]) for r in category_rows if r["id"]==category),1) if category else 1
            if featured and not active: row_errors.append("Featured=Yes nhưng Available=No")
            if featured and not cat_active: row_errors.append("Featured=Yes nhưng danh mục đang ẩn")
            if row_errors:
                errors.append({"row":excel_row_num,"message":"; ".join(row_errors)})
                continue
            if not image and existing_row:
                image=str(existing_row["image"] or "")
            if not image:
                # New rows without Image keep a generated placeholder SVG path.
                image=""
            action="update" if existing_row else "create"
            parsed.append({"row":excel_row_num,"name":name,"category":category,"price":price,
                           "description":_clean_import_value(raw["Description"]),"badge":_clean_import_value(raw["Badge"]),
                           "active":active,"featured":featured,"image":image,"action":action,
                           "existing_id":str(existing_row["id"]) if existing_row else "",
                           "sort_order":int(existing_row["sort_order"]) if existing_row else 0})
            if not image and not existing_row:
                warnings.append({"row":excel_row_num,"message":"Không có Image; món mới sẽ dùng ảnh placeholder mặc định."})
        wb.close()
        return parsed,errors,warnings
    except ValueError:
        raise
    except Exception as exc:
        raise ValueError(f"Không đọc được file Excel: {exc}") from exc


@app.get("/admin/menu/import-template")
def menu_import_template():
    if not session.get("admin"):
        return redirect("/admin")
    if load_workbook is None:
        return jsonify(ok=False,message="Chưa cài openpyxl. Hãy chạy: py -m pip install openpyxl"),500
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl import Workbook
    wb=Workbook()
    ws=wb.active; ws.title="Menu"
    headers=IMPORT_COLUMNS
    ws.append(headers)
    examples=[
        ["pho","Phở Đặc Biệt",16.90,"Beef pho with rare steak, brisket, tendon and meatball.","Best Seller","Yes","Yes","pho-dac-biet_d3d371fb4ca8.png"],
        ["banh-mi","Bánh Mì Đặc Biệt",9.90,"Vietnamese cold cuts, pâté and fresh vegetables in a crispy baguette.","Signature","Yes","No","banh-mi-dac-biet_9353a2073b16.png"],
        ["bun","Bún Thịt/Gà/Tôm Nướng",12.90,"Rice vermicelli with grilled meat, chicken or shrimp and fresh vegetables.","","Yes","No",""],
        ["com","Cơm Tay Cầm",13.90,"Hot clay-pot fried rice with shrimp, chicken, beef and vegetables.","","Yes","No",""],
        ["trang-mieng","Chè Ba Màu",6.90,"Vietnamese three-color dessert with jelly and coconut milk.","","Yes","No","che-ba-mau_che3mauupweb.png"],
    ]
    for row in examples: ws.append(row)
    widths=[24,34,14,62,18,14,12,38]
    for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
    header_fill=PatternFill("solid",fgColor="253E2E")
    header_font=Font(color="FFFFFF",bold=True)
    thin=Side(style="thin",color="D9D0C3")
    for cell in ws[1]:
        cell.fill=header_fill; cell.font=header_font; cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=Border(bottom=thin)
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:H{max(2,ws.max_row)}"
    for row in ws.iter_rows(min_row=2):
        row[2].number_format='$0.00'
        for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
    dv=DataValidation(type="list",formula1='"Yes,No"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"F2:G{MAX_IMPORT_ROWS+1}")

    guide=wb.create_sheet("Guide")
    guide_rows=[
        ["K&K — HƯỚNG DẪN IMPORT MENU"],
        ["Cột","Bắt buộc","Cách nhập"],
        ["Category","Có","Nhập ID danh mục (ví dụ pho) hoặc đúng tên danh mục trong Admin. Có thể dùng dạng song ngữ: Phở | Noodle Soup."],
        ["Name","Có","Tên món. Nếu Category + Name đã tồn tại thì hệ thống CẬP NHẬT món đó; nếu chưa có thì TẠO món mới."],
        ["Price_USD","Có","Giá USD, ví dụ 16.90. Không nhập dấu đ hoặc VND."],
        ["Description","Không","Mô tả món."],
        ["Badge","Không","Ví dụ: Best Seller, Signature. Để trống nếu không dùng."],
        ["Available","Không","Yes = đang bán; No = tạm ẩn. Mặc định Yes nếu để trống."],
        ["Featured","Không","Yes = Món nổi bật; No = không nổi bật. Tối đa 8 món nổi bật và món phải đang bán."],
        ["Image","Không","Tên file ảnh đã có trong uploads hoặc static/dishes. Có thể để trống; món mới sẽ dùng ảnh placeholder mặc định."],
        ["Lưu ý an toàn","","Hệ thống tạo backup database trước khi Import. Nếu Import lỗi, giao dịch database được rollback."],
        ["Giới hạn","","Tối đa 500 dòng món cho mỗi lần Import. Chỉ hỗ trợ .xlsx."],
    ]
    for r in guide_rows: guide.append(r)
    guide.column_dimensions["A"].width=24; guide.column_dimensions["B"].width=14; guide.column_dimensions["C"].width=105
    guide.merge_cells("A1:C1"); guide["A1"].font=Font(size=16,bold=True,color="FFFFFF"); guide["A1"].fill=header_fill
    for cell in guide[2]: cell.font=Font(bold=True); cell.fill=PatternFill("solid",fgColor="F1EADF")
    for row in guide.iter_rows():
        for cell in row: cell.alignment=Alignment(vertical="top",wrap_text=True)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out,as_attachment=True,download_name="KK_Menu_Import_Template.xlsx",mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.post("/admin/menu/import/preview")
def menu_import_preview():
    if not session.get("admin"): return jsonify(ok=False,message="Phiên đăng nhập đã hết hạn."),401
    _cleanup_import_staging()
    f=request.files.get("excel_file")
    if not f or not f.filename:
        return jsonify(ok=False,message="Vui lòng chọn file Excel .xlsx."),400
    try:
        parsed,errors,warnings=_validate_import_workbook(f)
    except ValueError as exc:
        return jsonify(ok=False,message=str(exc)),400
    if not parsed and not errors:
        return jsonify(ok=False,message="File không có dòng dữ liệu hợp lệ."),400
    token=secrets.token_urlsafe(24)
    owner=str(session.get("admin_username") or "admin")
    staging=os.path.join(IMPORT_DIR,token+".json")
    payload={"owner":owner,"created_at":time.time(),"rows":parsed}
    try:
        with open(staging,"w",encoding="utf-8") as out: json.dump(payload,out,ensure_ascii=False)
    except OSError:
        return jsonify(ok=False,message="Không thể tạo phiên import tạm thời."),500
    return jsonify(ok=True,token=token,total=len(parsed)+len(errors),valid=len(parsed),errors=errors,warnings=warnings,
                    rows=parsed[:100],message="Đã kiểm tra file. Hãy xem trước và xác nhận import.")


@app.post("/admin/menu/import/commit")
def menu_import_commit():
    if not session.get("admin"): return jsonify(ok=False,message="Phiên đăng nhập đã hết hạn."),401
    token=(request.form.get("token") or "").strip()
    if not re.fullmatch(r"[A-Za-z0-9_-]{20,100}",token):
        return jsonify(ok=False,message="Phiên import không hợp lệ hoặc đã hết hạn."),400
    staging=os.path.join(IMPORT_DIR,token+".json")
    if not os.path.isfile(staging):
        return jsonify(ok=False,message="Phiên import đã hết hạn. Vui lòng chọn lại file Excel."),400
    try:
        with open(staging,"r",encoding="utf-8") as src: payload=json.load(src)
    except Exception:
        return jsonify(ok=False,message="Không đọc được dữ liệu import tạm thời."),400
    owner=str(session.get("admin_username") or "admin")
    if payload.get("owner")!=owner or time.time()-float(payload.get("created_at",0))>1800:
        try: os.remove(staging)
        except OSError: pass
        return jsonify(ok=False,message="Phiên import không còn hợp lệ. Vui lòng chọn lại file Excel."),400
    rows=payload.get("rows") or []
    if not rows: return jsonify(ok=False,message="Không có dữ liệu để import."),400
    if len(rows)>MAX_IMPORT_ROWS: return jsonify(ok=False,message=f"Vượt giới hạn {MAX_IMPORT_ROWS} món mỗi lần import."),400

    backup=backup_database(force=True)
    if not backup:
        return jsonify(ok=False,message="Không thể tạo backup trước khi import. Import đã bị hủy để bảo vệ dữ liệu."),500
    c=None; created=updated=0; featured_count=0; old_uploads_to_remove=[]
    try:
        c=conn(); c.execute("BEGIN")
        max_sort=c.execute("SELECT COALESCE(MAX(sort_order),-1) n FROM menu_items").fetchone()["n"]
        for item in rows:
            category=item["category"]; name=item["name"]
            existing=c.execute("SELECT id,sort_order,image FROM menu_items WHERE category=? AND lower(trim(name))=lower(trim(?)) ORDER BY id LIMIT 2",(category,name)).fetchall()
            if len(existing)>1:
                raise ValueError(f'Dòng {item["row"]}: database có nhiều món trùng "{name}" trong danh mục.')
            if existing:
                item_id=existing[0]["id"]; sort_order=int(existing[0]["sort_order"])
                old_image=str(existing[0]["image"] or "")
                image=item["image"] or old_image
                if not image: image=f"/static/dishes/{item_id}.svg"
                if old_image.startswith("/uploads/") and image != old_image:
                    old_uploads_to_remove.append(old_image)
                c.execute("""UPDATE menu_items SET name=?,category=?,price=?,description=?,image=?,badge=?,active=?,featured=?,sort_order=? WHERE id=?""",
                          (name,category,int(item["price"]),item["description"],image,item["badge"],int(item["active"]),int(item["featured"]),sort_order,item_id))
                updated+=1
            else:
                max_sort+=1; item_id=secrets.token_hex(8)
                image=item["image"] or f"/static/dishes/{item_id}.svg"
                c.execute("""INSERT INTO menu_items(id,name,category,price,description,image,badge,active,featured,sort_order)
                             VALUES(?,?,?,?,?,?,?,?,?,?)""",
                          (item_id,name,category,int(item["price"]),item["description"],image,item["badge"],int(item["active"]),int(item["featured"]),max_sort))
                created+=1
            if item["featured"]: featured_count+=1
        # Enforce the same homepage limits as manual Admin controls.
        c.execute("""UPDATE menu_items SET featured=0 WHERE id IN (
            SELECT m.id FROM menu_items m LEFT JOIN categories cat ON cat.id=m.category
            WHERE m.featured=1 AND (m.active<>1 OR cat.id IS NULL OR cat.active<>1)
        )""")
        count=c.execute("SELECT COUNT(*) n FROM menu_items m JOIN categories cat ON cat.id=m.category WHERE m.featured=1 AND m.active=1 AND cat.active=1").fetchone()["n"]
        if count>8:
            raise ValueError("Import tạo ra hơn 8 món nổi bật. Hãy giảm Featured=Yes xuống tối đa 8 món rồi thử lại.")
        c.commit()
    except Exception as exc:
        if c:
            try: c.rollback()
            except Exception: pass
        logger.exception("Menu Excel import failed")
        return jsonify(ok=False,message=str(exc) if isinstance(exc,ValueError) else "Import thất bại. Dữ liệu cũ vẫn được giữ nguyên."),400 if isinstance(exc,ValueError) else 500
    finally:
        if c:
            c.close()
        try: os.remove(staging)
        except OSError: pass
    for old_image in old_uploads_to_remove:
        try:
            old_file=os.path.join(UP,os.path.basename(old_image))
            if os.path.isfile(old_file): os.remove(old_file)
        except OSError:
            pass
    audit_log("IMPORT","menu_item","bulk",{"created":created,"updated":updated,"rows":len(rows),"backup":os.path.basename(backup)})
    return jsonify(ok=True,message=f"Import thành công {len(rows)} món.",created=created,updated=updated,backup=os.path.basename(backup),refresh=True)

@app.post("/admin/menu/add")
def menu_add():
    if not session.get("admin"): return jsonify(ok=False),401
    item_id=secrets.token_hex(8)
    try:
        save_menu_item(item_id)
    except ValueError as e:
        r=admin_result(str(e), refresh=False, status=400)
        return r if r is not None else redirect("/admin#menu")
    audit_log("CREATE","menu_item",item_id,{"name":request.form.get("name","").strip()})
    r=admin_result("Đã thêm món mới.", refresh=True)
    return r if r is not None else redirect("/admin#menu")

@app.post("/admin/menu/save")
def menu_save():
    if not session.get("admin"): return jsonify(ok=False),401
    item_id=(request.form.get("id") or "").strip()
    if not item_id:
        r=admin_result("Thiếu mã món.", refresh=False)
        return r if r is not None else redirect("/admin#menu")
    try:
        save_menu_item(item_id)
    except ValueError as e:
        r=admin_result(str(e), refresh=False, status=400)
        return r if r is not None else redirect("/admin#menu")
    audit_log("UPDATE","menu_item",item_id,{"name":request.form.get("name","").strip(),"price":request.form.get("price","")})
    r=admin_result("Đã lưu món ăn.", refresh=True)
    return r if r is not None else redirect("/admin#menu")

@app.post("/admin/menu/delete/<item_id>")
def menu_delete(item_id):
    if not session.get("admin"): return jsonify(ok=False),401
    c=conn(); row=c.execute("SELECT image FROM menu_items WHERE id=?",(item_id,)).fetchone()
    if not row:
        c.close()
        return jsonify(ok=False,message="Không tìm thấy món ăn."),404
    c.execute("DELETE FROM menu_items WHERE id=?",(item_id,)); c.commit()
    c.close()
    audit_log("DELETE","menu_item",item_id,{"name":row["image"]})
    if row["image"].startswith("/uploads/"):
        try: os.remove(os.path.join(UP,os.path.basename(row["image"])))
        except OSError: pass
    r=admin_result("Đã xóa món ăn.", refresh=True)
    return r if r is not None else redirect("/admin#menu")

@app.post("/admin/menu/featured/<item_id>")
def menu_featured(item_id):
    if not session.get("admin"): return jsonify(ok=False),401
    data=request.json or {}
    featured=1 if data.get("featured") else 0
    c=conn()
    row=c.execute("""SELECT m.id,m.active,c.active AS category_active
        FROM menu_items m JOIN categories c ON c.id=m.category WHERE m.id=?""",(item_id,)).fetchone()
    if not row:
        c.close(); return jsonify(ok=False,message="Không tìm thấy món ăn."),404
    if featured and (not row["active"] or not row["category_active"]):
        c.close(); return jsonify(ok=False,message="Chỉ có thể đưa món đang bán thuộc danh mục đang hiển thị vào Món nổi bật."),400
    if featured:
        count=c.execute("""SELECT COUNT(*) n FROM menu_items m
            JOIN categories cat ON cat.id=m.category
            WHERE m.featured=1 AND m.active=1 AND cat.active=1 AND m.id<>?""",(item_id,)).fetchone()["n"]
        if count >= 8:
            c.close(); return jsonify(ok=False,message="Chỉ được chọn tối đa 8 món nổi bật trên trang chủ."),400
    c.execute("UPDATE menu_items SET featured=? WHERE id=?",(featured,item_id))
    c.commit(); c.close()
    audit_log("FEATURED","menu_item",item_id,{"featured":featured})
    return jsonify(ok=True,message="Đã cập nhật món nổi bật.")

@app.post("/admin/menu/reorder")
def menu_reorder():
    if not session.get("admin"): return jsonify(ok=False),401
    data=request.json or {}
    ids=data.get("ids") or []
    featured=data.get("featured") or {}
    if not isinstance(ids,list) or not isinstance(featured,dict):
        return jsonify(ok=False,message="Dữ liệu sắp xếp không hợp lệ."),400
    selected=[item_id for item_id in ids if featured.get(item_id)]
    if len(selected)>8:
        return jsonify(ok=False,message="Chỉ được chọn tối đa 8 món nổi bật."),400
    c=conn()
    valid={r["id"] for r in c.execute("""SELECT m.id FROM menu_items m
        JOIN categories cat ON cat.id=m.category
        WHERE m.active=1 AND cat.active=1""").fetchall()}
    if len(ids)!=len(set(ids)) or set(ids)!=valid:
        c.close(); return jsonify(ok=False,message="Danh sách món đã thay đổi. Vui lòng tải lại trang Admin rồi thử lại."),400
    if any(item_id not in valid for item_id in ids):
        c.close(); return jsonify(ok=False,message="Danh sách món không hợp lệ."),400
    for i,item_id in enumerate(ids):
        c.execute("UPDATE menu_items SET sort_order=? WHERE id=?",(i,item_id))
        c.execute("UPDATE menu_items SET featured=? WHERE id=?",(1 if featured.get(item_id) else 0,item_id))
    # Items hidden by an inactive category cannot remain featured.
    c.execute("""UPDATE menu_items SET featured=0 WHERE id IN (
        SELECT m.id FROM menu_items m
        LEFT JOIN categories cat ON cat.id=m.category
        WHERE m.featured=1 AND (m.active<>1 OR cat.id IS NULL OR cat.active<>1)
    )""")
    c.commit(); c.close()
    audit_log("REORDER_FEATURED","menu_item","bulk",{"count":len(selected)})
    return jsonify(ok=True,message="Đã lưu thứ tự và món nổi bật.")

@app.post("/admin/category/add")
def category_add():
    if not session.get("admin"): return jsonify(ok=False),401
    name=(request.form.get("name") or "").strip()
    cid=secure_filename(name).lower().replace("_","-")
    if not cid or not name:
        r=admin_result("Tên danh mục không hợp lệ.", refresh=False, status=400)
        return r if r is not None else redirect("/admin#categories")
    c=conn(); exists=c.execute("SELECT 1 FROM categories WHERE id=?",(cid,)).fetchone()
    if not exists:
        n=c.execute("SELECT COALESCE(MAX(sort_order),-1)+1 n FROM categories").fetchone()["n"]
        c.execute("INSERT INTO categories VALUES(?,?,?,1)",(cid,name,n)); c.commit()
    else:
        c.close(); r=admin_result("Danh mục đã tồn tại.", refresh=False, status=409); return r if r is not None else redirect("/admin#categories")
    c.close(); audit_log("CREATE","category",cid,{"name":name})
    r=admin_result("Đã thêm danh mục."); return r if r is not None else redirect("/admin#categories")

@app.post("/admin/category/save")
def category_save():
    if not session.get("admin"): return jsonify(ok=False),401
    cid=(request.form.get("id") or "").strip(); name=(request.form.get("name") or "").strip()
    if not cid or not name:
        r=admin_result("Tên danh mục không được để trống.", refresh=False, status=400)
        return r if r is not None else redirect("/admin#categories")
    c=conn(); cur=c.execute("UPDATE categories SET name=? WHERE id=?",(name,cid)); c.commit(); c.close()
    if cur.rowcount == 0:
        r=admin_result("Không tìm thấy danh mục.", refresh=False, status=404)
        return r if r is not None else redirect("/admin#categories")
    audit_log("UPDATE","category",cid,{"name":name})
    r=admin_result("Đã cập nhật danh mục."); return r if r is not None else redirect("/admin#categories")

@app.post("/admin/category/delete/<cid>")
def category_delete(cid):
    if not session.get("admin"): return jsonify(ok=False),401
    c=conn(); used=c.execute("SELECT COUNT(*) n FROM menu_items WHERE category=?",(cid,)).fetchone()["n"]
    if used:
        c.close(); r=admin_result(f"Danh mục đang có {used} món. Hãy chuyển hoặc xóa các món trước.", refresh=False, status=409); return r if r is not None else redirect("/admin#categories")
    cur=c.execute("DELETE FROM categories WHERE id=?",(cid,)); c.commit(); c.close()
    if cur.rowcount==0:
        r=admin_result("Không tìm thấy danh mục.", refresh=False, status=404)
        return r if r is not None else redirect("/admin#categories")
    audit_log("DELETE","category",cid,{})
    r=admin_result("Đã xóa danh mục."); return r if r is not None else redirect("/admin#categories")

@app.post("/admin/category/toggle/<cid>")
def category_toggle(cid):
    if not session.get("admin"): return jsonify(ok=False),401
    c=conn(); cur=c.execute("UPDATE categories SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE id=?",(cid,)); c.commit(); c.close()
    if cur.rowcount==0:
        r=admin_result("Không tìm thấy danh mục.", refresh=False, status=404)
        return r if r is not None else redirect("/admin#categories")
    audit_log("TOGGLE","category",cid,{})
    r=admin_result("Đã cập nhật trạng thái danh mục."); return r if r is not None else redirect("/admin#categories")

@app.post("/admin/category/move/<cid>/<direction>")
def category_move(cid,direction):
    if not session.get("admin"): return jsonify(ok=False,message="Phiên đăng nhập đã hết hạn."),401
    if direction not in ("up","down"):
        return jsonify(ok=False,message="Hướng sắp xếp không hợp lệ."),400
    c=conn(); rows=c.execute("SELECT id,sort_order FROM categories ORDER BY sort_order,id").fetchall(); ids=[r["id"] for r in rows]
    if cid in ids:
        i=ids.index(cid); j=i-1 if direction=="up" else i+1
        if 0<=j<len(ids):
            a,b=ids[i],ids[j]; sa=rows[i]["sort_order"]; sb=rows[j]["sort_order"]
            c.execute("UPDATE categories SET sort_order=? WHERE id=?",(sb,a)); c.execute("UPDATE categories SET sort_order=? WHERE id=?",(sa,b)); c.commit()
    c.close(); audit_log("MOVE","category",cid,{"direction":direction})
    r=admin_result("Đã cập nhật thứ tự danh mục."); return r if r is not None else redirect("/admin#categories")

@app.post("/admin/voucher")
def voucher_save():
    if not session.get("admin"): return jsonify(ok=False),401
    code=(request.form.get("code") or "").strip().upper()
    old_code=(request.form.get("old_code") or "").strip().upper()
    if not code:
        r=admin_result("Vui lòng nhập mã voucher.", refresh=False, status=400); return r if r is not None else redirect("/admin#voucher")
    if len(code)>40 or not re.fullmatch(r"[A-Z0-9_-]+",code):
        r=admin_result("Mã voucher chỉ được gồm A-Z, 0-9, dấu gạch ngang hoặc gạch dưới.", refresh=False, status=400)
        return r if r is not None else redirect("/admin#voucher")
    vtype=request.form.get("type","percent")
    if vtype not in VOUCHER_TYPES:
        r=admin_result("Loại voucher không hợp lệ.", refresh=False, status=400); return r if r is not None else redirect("/admin#voucher")
    try:
        if vtype == "fixed":
            value=parse_money_cents(request.form.get("value"))
        else:
            value=int(request.form.get("value") or 0)
        min_order=parse_money_cents(request.form.get("min_order"))
    except ValueError:
        r=admin_result("Giá trị voucher phải là số tiền hợp lệ.", refresh=False, status=400); return r if r is not None else redirect("/admin#voucher")
    if value<0 or min_order<0 or (vtype=="percent" and value>100):
        r=admin_result("Giá trị voucher không hợp lệ.", refresh=False, status=400); return r if r is not None else redirect("/admin#voucher")
    start_date=(request.form.get("start_date") or "").strip() or None
    end_date=(request.form.get("end_date") or "").strip() or None
    for date_value in (start_date,end_date):
        if date_value:
            try: datetime.strptime(date_value,"%Y-%m-%d")
            except ValueError:
                r=admin_result("Ngày voucher không hợp lệ.", refresh=False, status=400)
                return r if r is not None else redirect("/admin#voucher")
    if start_date and end_date and start_date>end_date:
        r=admin_result("Ngày kết thúc phải từ ngày áp dụng trở đi.", refresh=False, status=400); return r if r is not None else redirect("/admin#voucher")
    c=conn()
    try:
        if old_code and old_code!=code:
            if c.execute("SELECT 1 FROM vouchers WHERE code=?",(code,)).fetchone():
                c.close(); r=admin_result("Mã voucher mới đã tồn tại.", refresh=False, status=409)
                return r if r is not None else redirect("/admin#voucher")
            c.execute("UPDATE vouchers SET code=?,type=?,value=?,min_order=?,start_date=?,end_date=? WHERE code=?",
                      (code,vtype,value,min_order,start_date,end_date,old_code))
            if c.total_changes==0:
                c.close(); r=admin_result("Không tìm thấy voucher cần sửa.", refresh=False, status=404)
                return r if r is not None else redirect("/admin#voucher")
        else:
            existing=c.execute("SELECT active FROM vouchers WHERE code=?",(code,)).fetchone()
            active=int(existing["active"]) if existing else 1
            c.execute("""INSERT INTO vouchers
                (code,type,value,min_order,active,start_date,end_date)
                VALUES(?,?,?,?,?,?,?)
                ON CONFLICT(code) DO UPDATE SET
                  type=excluded.type,value=excluded.value,min_order=excluded.min_order,
                  start_date=excluded.start_date,end_date=excluded.end_date""",
                (code,vtype,value,min_order,active,start_date,end_date))
        c.commit()
    finally:
        c.close()
    audit_log("UPSERT","voucher",code,{"type":vtype,"value":value,"min_order":min_order})
    r=admin_result("Đã lưu voucher."); return r if r is not None else redirect("/admin#voucher")


@app.post("/admin/voucher/delete/<code>")
def voucher_delete(code):
    if not session.get("admin"): return jsonify(ok=False),401
    c=conn(); cur=c.execute("DELETE FROM vouchers WHERE code=?",(code,)); c.commit(); c.close()
    if cur.rowcount==0:
        r=admin_result("Không tìm thấy voucher.", refresh=False, status=404)
        return r if r is not None else redirect("/admin#voucher")
    audit_log("DELETE","voucher",code,{})
    r=admin_result("Đã xóa voucher."); return r if r is not None else redirect("/admin#voucher")

@app.post("/admin/voucher/toggle/<code>")
def voucher_toggle(code):
    if not session.get("admin"): return jsonify(ok=False),401
    c=conn(); cur=c.execute("UPDATE vouchers SET active=CASE active WHEN 1 THEN 0 ELSE 1 END WHERE code=?",(code,))
    c.commit(); c.close()
    if cur.rowcount==0:
        r=admin_result("Không tìm thấy voucher.", refresh=False, status=404)
        return r if r is not None else redirect("/admin#voucher")
    audit_log("TOGGLE","voucher",code,{})
    r=admin_result("Đã cập nhật trạng thái voucher."); return r if r is not None else redirect("/admin#voucher")

@app.post("/admin/settings")
def settings_save():
    if not session.get("admin"):
        return jsonify(ok=False, message="Phiên đăng nhập đã hết hạn."), 401

    # Lấy dữ liệu form trước khi xử lý file để hoạt động ổn định với AJAX FormData.
    location_image_keys=("location_main_image","location_side_image_1","location_side_image_2")
    values={k:(request.form.get(k) or "").strip() for k in DEFAULT_SETTINGS if k not in ("hero_image", "logo_image", "online_order_enabled", *location_image_keys)}
    values["maps_url"]=(request.form.get("maps_url") or "").strip()
    # Checkbox gửi kèm hidden value 0 để khi bỏ chọn vẫn lưu được trạng thái tắt.
    # Dùng getlist() vì form có cả hidden input và checkbox cùng name.
    order_values=request.form.getlist("online_order_enabled")
    values["online_order_enabled"]="1" if any(str(v).strip().lower() in ("1","true","yes","on") for v in order_values) else "0"
    current=request.form.get("hero_image") or DEFAULT_SETTINGS["hero_image"]
    logo_current=request.form.get("logo_image") or DEFAULT_SETTINGS.get("logo_image","")
    location_current={k:(request.form.get(k) or DEFAULT_SETTINGS[k]) for k in location_image_keys}

    if not values["restaurant_name"]:
        r=admin_result("Tên nhà hàng không được để trống.", refresh=False, status=400)
        return r if r is not None else redirect("/admin#settings")
    if values["maps_url"] and not re.match(r"^https?://", values["maps_url"], re.I):
        r=admin_result("Google Maps URL phải bắt đầu bằng http:// hoặc https://.", refresh=False, status=400)
        return r if r is not None else redirect("/admin#settings")
    if values["google_review_url"] and not re.match(r"^https?://", values["google_review_url"], re.I):
        r=admin_result("Link đánh giá Google phải bắt đầu bằng http:// hoặc https://.", refresh=False, status=400)
        return r if r is not None else redirect("/admin#settings")

    banner=request.files.get("hero_image_file")
    logo=request.files.get("logo_image_file")
    location_files={k:request.files.get(f"{k}_file") for k in location_image_keys}
    old_banner=current
    old_logo=logo_current
    old_location=dict(location_current)
    saved_files=[]
    new_banner=current
    new_logo=logo_current
    new_location=dict(location_current)

    # Validate tất cả file trước khi ghi DB để tránh trạng thái lưu dở dang.
    uploads=[]
    for file_obj, label in ((banner, "banner"), (logo, "logo"), *[(location_files[k], k) for k in location_image_keys]):
        if not file_obj or not file_obj.filename:
            continue
        original=secure_filename(file_obj.filename)
        ext=original.rsplit(".",1)[-1].lower() if "." in original else ""
        if ext not in IMAGE_EXTENSIONS:
            label_names={"banner":"Ảnh banner","logo":"Ảnh logo","location_main_image":"Ảnh chính Địa điểm","location_side_image_1":"Ảnh phụ 1 Địa điểm","location_side_image_2":"Ảnh phụ 2 Địa điểm"}
            message=label_names.get(label,"Ảnh")
            r=admin_result(f"{message} phải có định dạng JPG, JPEG, PNG hoặc WEBP.", refresh=False, status=400)
            return r if r is not None else redirect("/admin#settings")
        prefix_map={"banner":"hero_banner","logo":"site_logo","location_main_image":"location_main","location_side_image_1":"location_side_1","location_side_image_2":"location_side_2"}
        prefix=prefix_map.get(label,label)
        filename=f"{prefix}_{secrets.token_hex(8)}.{ext}"
        path=os.path.join(UP, filename)
        uploads.append((file_obj, label, path, "/uploads/"+filename))

    try:
        for file_obj, label, path, public_path in uploads:
            ext=path.rsplit(".",1)[-1].lower()
            save_validated_image(file_obj, path, ext)
            saved_files.append(path)
            if label=="banner":
                new_banner=public_path
            elif label=="logo":
                new_logo=public_path
            elif label in new_location:
                new_location[label]=public_path

        c=conn()
        for k,v in values.items():
            c.execute("INSERT OR REPLACE INTO settings VALUES(?,?)", (k,v))
        c.execute("INSERT OR REPLACE INTO settings VALUES(?,?)", ("hero_image",new_banner))
        c.execute("INSERT OR REPLACE INTO settings VALUES(?,?)", ("logo_image",new_logo))
        for k,v in new_location.items():
            c.execute("INSERT OR REPLACE INTO settings VALUES(?,?)", (k,v))
        c.commit()
        c.close()
    except Exception:
        try:
            c.close()
        except Exception:
            pass
        for path in saved_files:
            try: os.remove(path)
            except OSError: pass
        r=admin_result("Không thể lưu cài đặt hoặc hình ảnh. Vui lòng thử lại.", refresh=False, status=500)
        return r if r is not None else redirect("/admin#settings")

    # Chỉ xóa file cũ sau khi DB đã commit thành công.
    for old_path, new_path in [(old_banner,new_banner),(old_logo,new_logo), *[(old_location[k],new_location[k]) for k in location_image_keys]]:
        if old_path and old_path.startswith("/uploads/") and old_path != new_path:
            try:
                old_file=os.path.join(UP, os.path.basename(old_path))
                if os.path.isfile(old_file):
                    os.remove(old_file)
            except OSError:
                pass

    audit_log("UPDATE","settings","restaurant",{"uploaded_banner":bool(uploads and any(x[1]=="banner" for x in uploads)),
                                               "uploaded_logo":bool(uploads and any(x[1]=="logo" for x in uploads)),
                                               "uploaded_location_images":bool(uploads and any(x[1] in location_image_keys for x in uploads))})
    r=admin_result("Đã lưu thông tin nhà hàng, logo, banner và Địa điểm trang chủ.")
    return r if r is not None else redirect("/admin#settings")


@app.errorhandler(Exception)
def handle_unexpected_error(error):
    if isinstance(error, HTTPException):
        return error
    # Keep production users away from Werkzeug tracebacks while logging full details.
    logger.exception("Unhandled exception on %s %s", request.method, request.path)
    if is_ajax_request() or request.path.startswith("/api/"):
        return jsonify(ok=False,message="Đã xảy ra lỗi máy chủ. Vui lòng thử lại sau."),500
    return render_template("error.html"),500

@app.get("/healthz")
def healthz():
    c=None
    try:
        c=conn()
        ok=c.execute("PRAGMA quick_check").fetchone()[0]=="ok"
        return jsonify(ok=ok), 200 if ok else 503
    except Exception:
        logger.exception("healthz failed")
        return jsonify(ok=False),503
    finally:
        if c:
            c.close()

@app.get("/admin/health")
def admin_health():
    if not session.get("admin"):
        return jsonify(ok=False,message="Unauthorized"),401
    checks={}
    c=None
    try:
        c=conn()
        checks["database_connection"]=True
        checks["sqlite_integrity"]=c.execute("PRAGMA integrity_check").fetchone()[0]=="ok"
        checks["foreign_keys"]=c.execute("PRAGMA foreign_keys").fetchone()[0]==1
        checks["journal_mode"]=str(c.execute("PRAGMA journal_mode").fetchone()[0]).lower()=="wal"
        required={"admin_account","menu_items","categories","orders","bookings","vouchers","settings","audit_logs","rate_limit_hits"}
        present={r["name"] for r in c.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
        checks["required_tables"]=required.issubset(present)
        required_indexes={"idx_orders_idempotency","idx_orders_status_created","idx_orders_phone_code",
                          "idx_bookings_date_status","idx_menu_category_active_order",
                          "idx_categories_active_order","idx_rate_limit_lookup","idx_audit_logs_created"}
        present_idx={r["name"] for r in c.execute("SELECT name FROM sqlite_master WHERE type='index'").fetchall()}
        checks["required_indexes"]=required_indexes.issubset(present_idx)
    except Exception:
        logger.exception("Health check failed")
        checks["database_connection"]=False
    finally:
        if c:
            c.close()
    backups=[os.path.join(BACKUP_DIR,x) for x in os.listdir(BACKUP_DIR) if x.startswith("kk_") and x.endswith(".db")]
    newest=max(backups,key=os.path.getmtime) if backups else None
    checks["backup_available"]=bool(newest and time.time()-os.path.getmtime(newest)<48*3600)
    checks["backup_file"]=os.path.basename(newest) if newest else None
    checks["https_configured"]=bool(FORCE_HTTPS and HSTS_ENABLED)
    checks["csp_enabled"]=bool(CSP_ENABLED)
    core_keys={"database_connection","sqlite_integrity","foreign_keys","journal_mode","required_tables","required_indexes"}
    checks["overall"]=all(checks.get(k,False) for k in core_keys)
    return jsonify(ok=checks["overall"],checks=checks)

@app.get("/admin/audit-logs")
def admin_audit_logs():
    if not session.get("admin"):
        return jsonify(ok=False,message="Unauthorized"),401
    try:
        limit=min(max(int(request.args.get("limit",100)),1),500)
    except (TypeError, ValueError):
        limit=100
    try:
        page=max(int(request.args.get("page",1)),1)
    except (TypeError, ValueError):
        page=1

    c=conn()
    try:
        total=int(c.execute("SELECT COUNT(*) n FROM audit_logs").fetchone()["n"])
        pages=max((total + limit - 1)//limit, 1)
        page=min(page,pages)
        offset=(page-1)*limit
        rows=[dict(r) for r in c.execute(
            """SELECT id,created_at,username,ip,action,entity_type,entity_id,details
               FROM audit_logs ORDER BY id DESC LIMIT ? OFFSET ?""",
            (limit,offset)).fetchall()]
    finally:
        c.close()
    return jsonify(ok=True,logs=rows,total=total,page=page,pages=pages,limit=limit)

@app.route("/uploads/<path:name>")
def uploads(name): return send_from_directory(UP,name)

@app.route("/robots.txt")
def robots():
    return app.response_class("User-agent: *\nAllow: /\nDisallow: /admin\nDisallow: /api/\nSitemap: "+request.url_root.rstrip("/")+"/sitemap.xml\n",mimetype="text/plain")

@app.route("/sitemap.xml")
def sitemap():
    h=request.url_root.rstrip("/")
    xml='<?xml version="1.0" encoding="UTF-8"?><urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'
    paths=["/","/menu"]
    if online_order_enabled():
        paths += ["/book","/track"]
    for path in paths:
        xml += f"<url><loc>{h}{path}</loc></url>"
    xml += "</urlset>"
    return app.response_class(xml,mimetype="application/xml")

if __name__=="__main__":
    app.run(host="0.0.0.0", port=5000, debug=False)
    
